import numpy as np

import openpyxl as xl
from openpyxl.utils import get_column_letter

def load_pft():
    filename = "PF-T.xlsm"
    sheetname_pft = "PF-T"
    sheetname_scenario = "Scenario"
    sheetname_pft_bas = "PF-T bas"
    wb = xl.load_workbook(filename=filename, read_only=False, data_only=True, keep_vba=True)
    ws_pft = wb[sheetname_pft]
    ws_scenario = wb[sheetname_scenario]
    ws_pft_bas = wb[sheetname_pft_bas]

    return wb, ws_pft, ws_scenario, ws_pft_bas


def dico_scenario(ws_scenario):
    dico = {}
    dico_nd = {"Liste des key ND"}
    first_line = 3
    last_line = 28
    col_sp = 5
    col_moa = 4
    col_typo = 18

    for line in range(first_line, last_line+1):
        sp = ws_scenario.cell(line, col_sp).value
        # la clé est "MOA-SP" s'il y a une SP, sinon il s'agit seulement de "MOA"
        if sp is None:
            key = str(ws_scenario.cell(line, col_moa).value)
            dico[key] = {}
            for column in range(9):
                dico[key][str(ws_scenario.cell(2, 9 + column).value)] \
                    = str(ws_scenario.cell(line, 9 + column).value or 0)

            if str(ws_scenario.cell(line, col_typo).value) == "ND":
                dico_nd.add(key)

        else:
            key = f"{ws_scenario.cell(line, col_moa).value}-{ws_scenario.cell(line, col_sp).value}"
            dico[key] = {}
            for column in range(9):
                dico[key][str(ws_scenario.cell(2, 9 + column).value)] \
                    = str(ws_scenario.cell(line, 9 + column).value or 0)

            if str(ws_scenario.cell(line, col_typo).value) == "ND":
                dico_nd.add(key)

    return dico, dico_nd
