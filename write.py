import numpy as np
import openpyxl as xl
from openpyxl.utils.cell import column_index_from_string
from datetime import datetime

from util import addYears


def use_scenario(sheet_pft, scenario, sheet_pft_bas, dico_nd):
    first_line = 13
    last_line = sheet_pft.max_row
    # today = datetime.now()
    # Date des données, je ne décale que les jalons présents après cette date
    today = datetime.strptime("02/05/22 10:00:00", "%d/%m/%y %H:%M:%S")

    col_typo = "B"
    pos_typo = column_index_from_string(col_typo)
    col_statut = "E"
    pos_statut = column_index_from_string(col_statut)
    col_moa = "DD"
    pos_moa = column_index_from_string(col_moa)
    col_prio = "SD"
    pos_prio = column_index_from_string(col_prio)
    col_sp = "DI"
    pos_sp = column_index_from_string(col_sp)

    col_consistances = "DJ"
    pos_consistances = int(column_index_from_string(col_consistances))

    col_ressources_bleues = "SP"
    pos_ressources_bleues = int(column_index_from_string(col_ressources_bleues))

    col_ressources_autre_di = "WH"
    pos_ressources_autre_di = int(column_index_from_string(col_ressources_autre_di))

    col_probable = "UL"
    pos_probable = int(column_index_from_string(col_probable))

    col_jalons = "SL"
    pos_jalons = int(column_index_from_string(col_jalons))

    col_prio_inter = "YB"
    pos_prio_inter = int(column_index_from_string(col_prio_inter))

    col_decalage_inter = "YC"
    pos_decalage_inter = int(column_index_from_string(col_decalage_inter))

    for line in range(first_line, last_line + 1):
        moa = str(sheet_pft.cell(line, pos_moa).value)
        sp = str(sheet_pft.cell(line, pos_sp).value)
        statut = str(sheet_pft.cell(line, pos_statut).value)
        typo = str(sheet_pft.cell(line, pos_typo).value)

        couple = f"{moa}-{sp}"
        key = None
        if moa in scenario:
            key = moa
        elif couple in scenario:
            key = couple

        # même si ma clé est vide, je force le décalage si un décalage inter est renseigné
        if sheet_pft.cell(line, pos_decalage_inter).value is None:
            decalage_inter_is_filled = False
        else:
            decalage_inter_is_filled = True

        if key is not None or decalage_inter_is_filled:

            # si la ligne du scénario n'est pas ND ou si la colonne B du portefeuille est Nd, je prends
            if key not in dico_nd or typo == "ND" or decalage_inter_is_filled:
                # si la pos_prio_inter est vide, je prends la prio classique
                if sheet_pft.cell(line, pos_prio_inter).value is None:
                    prio = str(sheet_pft.cell(line, pos_prio).value)
                else:
                    prio = str(sheet_pft.cell(line, pos_prio_inter).value)

                if not prio.isdigit() or prio == "0":
                    prio = "9"

                # j'applique le scénario à cette ligne
                # si la colonne pos_decalage_inter est vide, je prends le décalage classique
                if sheet_pft.cell(line, pos_decalage_inter).value is None:
                    decalage = int(scenario[key][prio])
                else:
                    decalage = int(sheet_pft.cell(line, pos_decalage_inter).value)

                sheet_pft_bas.cell(line, 1).value = decalage

                # consistances
                for num_consistance in range(32):
                    for annee in range(2, 2 + decalage):
                        # suppression des valeurs entre 2022 et le décalage
                        sheet_pft_bas.cell(line, pos_consistances + 12 * num_consistance + annee).value = None

                    for annee in range(2, 12 - decalage):
                        # decalage de la consistance
                        valeur_pft = str(sheet_pft.cell(line, pos_consistances + 12 * num_consistance + annee).value)
                        if valeur_pft == "None" or valeur_pft == "0" or valeur_pft == "":
                            sheet_pft_bas.cell(line,
                                               pos_consistances + 12 * num_consistance + annee + decalage).value = None
                        else:
                            valeur_pft = valeur_pft.replace(",", ".")
                            try:
                                sheet_pft_bas.cell(line,
                                                   pos_consistances + 12 * num_consistance + annee + decalage).value \
                                    = float(valeur_pft)
                            except:
                                sheet_pft_bas.cell(line,
                                                   pos_consistances + 12 * num_consistance + annee + decalage).value \
                                    = str(valeur_pft)

                # Ressources bleues
                for num_ressources in range(4):

                    for annee in range(2, 2 + decalage):
                        # suppression des valeurs entre 2022 et le décalage
                        sheet_pft_bas.cell(line, pos_ressources_bleues + 12 * num_ressources + annee).value = None

                    for annee in range(2, 12 - decalage):
                        valeur_pft = str(
                            sheet_pft.cell(line, pos_ressources_bleues + 12 * num_ressources + annee).value)
                        if valeur_pft == "None" or valeur_pft == "0" or valeur_pft == "":
                            sheet_pft_bas.cell(line,
                                               pos_ressources_bleues + 12 * num_ressources + annee + decalage).value = None
                        else:
                            sheet_pft_bas.cell(line,
                                               pos_ressources_bleues + 12 * num_ressources + annee + decalage).value = \
                                float(valeur_pft)

                # Ressources autre DI
                for annee in range(2, 2 + decalage):
                    # suppression des valeurs entre 2022 et le décalage
                    sheet_pft_bas.cell(line, pos_ressources_autre_di + annee).value = None

                for annee in range(2, 12 - decalage):
                    valeur_pft = str(sheet_pft.cell(line, pos_ressources_autre_di + annee).value)
                    if valeur_pft == "None" or valeur_pft == "0" or valeur_pft == "":
                        sheet_pft_bas.cell(line, pos_ressources_autre_di + annee + decalage).value = None
                    else:
                        sheet_pft_bas.cell(line, pos_ressources_autre_di + annee + decalage).value = float(valeur_pft)

                # Probable
                for annee in range(2, 2 + decalage):
                    # suppression des valeurs entre 2022 et le décalage
                    sheet_pft_bas.cell(line, pos_probable + annee).value = None

                for annee in range(2, 12 - decalage):
                    valeur_pft = str(sheet_pft.cell(line, pos_probable + annee).value)
                    if valeur_pft == "None" or valeur_pft == "0" or valeur_pft == "":
                        sheet_pft_bas.cell(line, pos_probable + annee + decalage).value = None
                    else:
                        sheet_pft_bas.cell(line, pos_probable + annee + decalage).value = float(valeur_pft)

                # Jalons
                if not (statut == "Jalons manquants" or statut == "---"):
                    for jalon in range(4):
                        valeur_pft = sheet_pft.cell(line, pos_jalons + jalon).value

                        if valeur_pft is None or str(valeur_pft) == "0" or str(valeur_pft) == "":
                            sheet_pft_bas.cell(line, pos_probable + jalon + decalage).value = None
                        elif isinstance(valeur_pft, str):
                            try:
                                valeur_pft = datetime.strptime(valeur_pft, '%d/%m/%Y')
                            except:
                                break
                            if valeur_pft > today:
                                # je décale seulement les jalons dans le futur
                                new_date = addYears(valeur_pft, decalage)
                                sheet_pft_bas.cell(line, pos_jalons + jalon).value = new_date
                        elif valeur_pft > today:
                            # je décale seulement les jalons dans le futur
                            new_date = addYears(valeur_pft, decalage)
                            sheet_pft_bas.cell(line, pos_jalons + jalon).value = new_date
                        else:
                            pass
        else:
            # pas de décalage, je ne fais rien
            pass
